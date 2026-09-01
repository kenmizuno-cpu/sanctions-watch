"""CSVマスターから配布用 Excel を生成する。

Excel はマスターではなく派生物。毎回ここで作り直す。
既存ファイルの仕様に合わせる:
  - シート1 は A〜G の7列のみ、ヘッダー名も既存と同一
  - 登録時間/更新時間は Unix epoch ミリ秒から JST の日時文字列へ変換して書く
"""
from __future__ import annotations

import argparse
import csv
from datetime import datetime, timezone, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .normalize import SRC_UNKNOWN

JST = timezone(timedelta(hours=9))
SHEET_MAIN = "Sheet1"
SHEET_LOG = "クリーンアップ履歴"
SHEET_INFO = "ビルド情報"

HEADERS = ["受取人名", "リスクタイプ", "有効なのか", "リスクレベル",
           "登録時間", "更新時間", "備考"]
WIDTHS = [60, 14, 11, 12, 15, 15, 52]

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=11)


def _jst(ms: str | int) -> str:
    try:
        return datetime.fromtimestamp(int(ms) / 1000, JST).strftime("%Y-%m-%d %H:%M:%S")
    except (ValueError, TypeError):
        return ""


def _style_header(ws, headers, widths) -> None:
    for i, h in enumerate(headers, start=1):
        c = ws.cell(1, i, h)
        c.fill, c.font = HEAD_FILL, HEAD_FONT
        c.alignment = Alignment(vertical="center")
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _read(path: Path) -> list[dict]:
    if not path.exists():
        return []
    with path.open(encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))


def build_workbook(master: list[dict], log: list[dict]) -> Workbook:
    wb = Workbook()

    # ---- Sheet1: 既存仕様の7列 -------------------------------------
    ws = wb.active
    ws.title = SHEET_MAIN
    _style_header(ws, HEADERS, WIDTHS)
    for r in master:
        ws.append([
            r["display_name"], r["risk_type"], r["status"], r["risk_level"],
            _jst(r["first_seen_ms"]), _jst(r["last_updated_ms"]), r["remark"],
        ])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for c in row:
            c.font = BODY_FONT
    # 登録時間/更新時間は JST の日時文字列として保持する
    for col in ("E", "F"):
        for c in ws[col][1:]:
            c.number_format = "@"

    # ---- クリーンアップ履歴 ----------------------------------------
    wl = wb.create_sheet(SHEET_LOG)
    lh = ["元Excel行", "種別", "元の値", "変更後", "理由"]
    _style_header(wl, lh, [11, 12, 58, 58, 62])
    for e in log:
        wl.append([e.get("行", ""), e.get("種別", ""),
                   e.get("元の値", ""), e.get("変更後", ""), e.get("理由", "")])
    for row in wl.iter_rows(min_row=2, max_row=wl.max_row):
        for c in row:
            c.font = BODY_FONT

    # ---- ビルド情報 -------------------------------------------------
    wi = wb.create_sheet(SHEET_INFO)
    _style_header(wi, ["項目", "値"], [30, 74])
    stats = {
        "生成日時 (JST)": datetime.now(JST).strftime("%Y-%m-%d %H:%M:%S"),
        "総行数": len(master),
        "うち有効": sum(1 for r in master if r["status"] == "有効"),
        "うち無効": sum(1 for r in master if r["status"] == "無効"),
        "クリーンアップ件数": len(log),
        "最古の登録": _jst(min((r["first_seen_ms"] for r in master), default=0)),
        "最新の更新": _jst(max((r["last_updated_ms"] for r in master), default=0)),
    }
    for src in sorted({s for r in master for s in r["sources"].split(";")
                       if s and s != SRC_UNKNOWN}):
        stats[f"出所: {src}"] = sum(1 for r in master
                                   if src in r["sources"].split(";"))
    for k, v in stats.items():
        wi.append([k, v])
    for row in wi.iter_rows(min_row=2, max_row=wi.max_row):
        for c in row:
            c.font = BODY_FONT
    return wb


ROOT = Path(__file__).resolve().parent.parent


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--indir", default=str(ROOT / "data" / "master"))
    ap.add_argument("--out", default=str(ROOT / "dist" / "black_receiver_name_all.xlsx"))
    args = ap.parse_args()

    d = Path(args.indir)
    master = _read(d / "master.csv")
    if not master:
        raise SystemExit(f"マスターが空: {d / 'master.csv'}")
    wb = build_workbook(master, _read(d / "cleanup_log.csv"))
    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"wrote {out} ({len(master)} 行)")


if __name__ == "__main__":
    main()
