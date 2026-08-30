"""差分CSVから社内取込用Excelを生成する。

出力するSheet1は既存マスターExcelと同じ7列。

重要:
OFAC Classic CSV運用中は「掲載終了候補」を検出しても、
master側ではまだ無効化していない。
そのような候補は社内取込ファイルから除外する。
"""
from __future__ import annotations

import argparse
import csv
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .normalize import match_key


ROOT = Path(__file__).resolve().parent.parent

MASTER = ROOT / "data" / "master" / "master.csv"
DIFF = ROOT / "data" / "diff" / "latest.csv"

JST = timezone(timedelta(hours=9))

HEADERS = [
    "受取人名",
    "リスクタイプ",
    "有効なのか",
    "リスクレベル",
    "登録時間",
    "更新時間",
    "備考",
]

WIDTHS = [60, 14, 11, 12, 15, 15, 52]

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(
    name="Arial",
    size=11,
    bold=True,
    color="FFFFFF",
)
BODY_FONT = Font(
    name="Arial",
    size=11,
)


def _read(path: Path) -> list[dict]:
    if not path.exists():
        return []

    with path.open(
        encoding="utf-8",
        newline="",
    ) as f:
        return list(csv.DictReader(f))


def _style_header(ws, headers, widths) -> None:
    for i, header in enumerate(headers, start=1):
        cell = ws.cell(1, i, header)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(vertical="center")

    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[
            get_column_letter(i)
        ].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = (
        f"A1:{get_column_letter(len(headers))}1"
    )


def _literal(cell, value) -> None:
    """外部データをExcel数式として評価させない。"""
    cell.value = "" if value is None else str(value)
    cell.data_type = "s"


def _append_row(ws, values) -> None:
    row_no = ws.max_row + 1

    for col_no, value in enumerate(values, start=1):
        cell = ws.cell(row_no, col_no)
        _literal(cell, value)
        cell.font = BODY_FONT


def select_actionable(
    master_rows: list[dict],
    diff_rows: list[dict],
) -> tuple[list[dict], int]:

    master = {
        row["match_key"]: row
        for row in master_rows
    }

    selected: list[dict] = []
    seen: set[str] = set()
    held = 0

    for diff in diff_rows:
        kind = (diff.get("種別") or "").strip()
        before = diff.get("変更前") or ""
        after = diff.get("変更後") or ""

        # OFACの「掲載終了候補」は
        # before == after のままmasterを変更していない。
        if kind == "掲載終了" and before == after:
            held += 1
            continue

        key = match_key(
            diff.get("受取人名", "")
        )

        if not key:
            continue

        row = master.get(key)

        if row is None:
            raise RuntimeError(
                "差分に対応するマスター行がありません: "
                + diff.get("受取人名", "")
            )

        # 同じ対象が複数ソースから同時更新されても
        # 社内取込は現在状態の1行だけ。
        if key in seen:
            continue

        seen.add(key)
        selected.append(row)

    return selected, held


def build_workbook(
    master_rows: list[dict],
    diff_rows: list[dict],
):
    selected, held = select_actionable(
        master_rows,
        diff_rows,
    )

    if not selected:
        return None, 0, held

    wb = Workbook()

    # 社内取込用
    ws = wb.active
    ws.title = "Sheet1"

    _style_header(
        ws,
        HEADERS,
        WIDTHS,
    )

    for row in selected:
        _append_row(
            ws,
            [
                row.get("display_name", ""),
                row.get("risk_type", ""),
                row.get("status", ""),
                row.get("risk_level", ""),
                row.get("first_seen_ms", ""),
                row.get("last_updated_ms", ""),
                row.get("remark", ""),
            ],
        )

    for col in ("E", "F"):
        for cell in ws[col][1:]:
            cell.number_format = "@"

    # 作業・監査用
    info = wb.create_sheet("差分情報")

    _style_header(
        info,
        ["項目", "値"],
        [34, 72],
    )

    values = [
        (
            "生成日時 (JST)",
            datetime.now(JST).strftime(
                "%Y-%m-%d %H:%M:%S"
            ),
        ),
        ("社内取込対象件数", len(selected)),
        ("差分CSV総件数", len(diff_rows)),
        ("掲載終了候補・保留件数", held),
    ]

    for key, value in values:
        _append_row(
            info,
            [key, value],
        )

    return wb, len(selected), held


def main() -> int:
    parser = argparse.ArgumentParser()

    parser.add_argument(
        "--master",
        default=str(MASTER),
    )

    parser.add_argument(
        "--diff",
        default=str(DIFF),
    )

    parser.add_argument(
        "--out",
        required=True,
    )

    args = parser.parse_args()

    master_rows = _read(
        Path(args.master)
    )

    diff_rows = _read(
        Path(args.diff)
    )

    if not master_rows:
        raise SystemExit(
            f"マスターが空です: {args.master}"
        )

    if not diff_rows:
        print("actionable_count=0")
        print("held_count=0")
        return 0

    wb, count, held = build_workbook(
        master_rows,
        diff_rows,
    )

    out = Path(args.out)

    if wb is None:
        if out.exists():
            out.unlink()

        print("社内取込対象差分なし")
        print("actionable_count=0")
        print(f"held_count={held}")
        return 0

    out.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    wb.save(out)

    print(f"wrote {out} ({count}件)")
    print(f"actionable_count={count}")
    print(f"held_count={held}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
